# zones_export_routes.py
# Rotas para:
#   /api/zones          — CRUD de zonas de cultivo
#   /api/relatorios/exportar — Exportação XLSX e PDF profissional

from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import db, User, Fazenda, Sensor, Zona, DadosIoT
from datetime import datetime, timedelta
from sqlalchemy import func
import io, math, os

zones_export_bp = Blueprint('zones_export', __name__)


# ═══════════════════════════════════════════════════════════════
#  HELPER — utilisador autenticado
# ═══════════════════════════════════════════════════════════════
def _get_user():
    uid = int(get_jwt_identity())
    return User.query.get(uid)


def _get_fazenda(user):
    if user and user.fazenda_id:
        return Fazenda.query.get(user.fazenda_id)
    return None


# ═══════════════════════════════════════════════════════════════
#  HELPER — criar zona padrão com base na localização da fazenda
# ═══════════════════════════════════════════════════════════════
def _criar_zona_padrao(user, fazenda):
    """
    Cria automaticamente uma zona de cultivo inicial após a primeira
    requisição, usando a latitude/longitude do último dado IoT da fazenda
    ou a localização textual da fazenda como nome.
    """
    try:
        # Descobrir lat/lng via último registo IoT da fazenda
        lat = lng = None
        localizacao_nome = fazenda.localizacao or "Zona Principal"

        sensores = Sensor.query.filter_by(fazenda_id=fazenda.id).all()
        device_ids = [s.nome for s in sensores]
        if device_ids:
            ultimo = (DadosIoT.query
                      .filter(DadosIoT.device_id.in_(device_ids))
                      .order_by(DadosIoT.timestamp.desc())
                      .first())
            if ultimo and ultimo.latitude and ultimo.longitude:
                lat = ultimo.latitude
                lng = ultimo.longitude
                # Gerar nome geográfico simplificado
                localizacao_nome = f"Zona {abs(lat):.2f}°{'N' if lat >= 0 else 'S'} {abs(lng):.2f}°{'E' if lng >= 0 else 'W'}"

        # Cultura principal da fazenda
        culturas = fazenda.get_culturas_list()
        cultura_principal = culturas[0].capitalize() if culturas else "Milho"
        tipo_map = {
            "cafe": "Frutíferas", "milho": "Milho", "soja": "Soja",
            "feijao": "Feijão", "feijão": "Feijão", "trigo": "Trigo",
            "arroz": "Arroz", "hortaliças": "Hortaliças"
        }
        tipo = tipo_map.get(culturas[0].lower() if culturas else "", "Milho")

        # Área total da fazenda como área inicial da zona
        area_ha = fazenda.hectares or 1.0

        zona = Zona(
            user_id=user.id,
            fazenda_id=fazenda.id,
            nome=localizacao_nome,
            cultura=cultura_principal,
            tipo=tipo,
            area_ha=area_ha,
            estagio="Crescimento",
            saude="Bom",
            acoes=f"Zona criada automaticamente. Lat: {lat:.4f}, Lng: {lng:.4f}" if lat else "Zona criada automaticamente a partir dos dados da fazenda."
        )
        db.session.add(zona)
        db.session.commit()
        return zona
    except Exception as e:
        print(f"[Zonas] Erro ao criar zona padrão: {e}")
        db.session.rollback()
        return None


# ═══════════════════════════════════════════════════════════════
#  GET /api/zones   — listar zonas do agricultor
# ═══════════════════════════════════════════════════════════════
@zones_export_bp.route("/api/zones", methods=["GET"])
@jwt_required()
def list_zones():
    user = _get_user()
    if not user:
        return jsonify({"erro": "Utilizador não encontrado"}), 404

    fazenda = _get_fazenda(user)
    fazenda_id = request.args.get("fazenda_id")

    q = Zona.query.filter_by(user_id=user.id)
    if fazenda_id:
        q = q.filter_by(fazenda_id=int(fazenda_id))

    zones = q.order_by(Zona.created_at.asc()).all()

    # Se não há zonas e o utilizador tem fazenda → criar zona padrão automaticamente
    if not zones and fazenda:
        zona_padrao = _criar_zona_padrao(user, fazenda)
        if zona_padrao:
            zones = [zona_padrao]

    return jsonify([z.to_dict() for z in zones]), 200


# ═══════════════════════════════════════════════════════════════
#  POST /api/zones  — criar zona
# ═══════════════════════════════════════════════════════════════
@zones_export_bp.route("/api/zones", methods=["POST"])
@jwt_required()
def create_zone():
    user = _get_user()
    if not user:
        return jsonify({"erro": "Utilizador não encontrado"}), 404

    d = request.get_json(silent=True) or {}
    nome    = (d.get("nome") or "").strip()
    cultura = (d.get("cultura") or "").strip()
    area    = d.get("area", 0)

    if not nome or not cultura:
        return jsonify({"erro": "Campos 'nome' e 'cultura' são obrigatórios"}), 400

    fazenda = _get_fazenda(user)
    zona = Zona(
        user_id=user.id,
        fazenda_id=fazenda.id if fazenda else None,
        nome=nome,
        cultura=cultura,
        tipo=d.get("tipo") or "Milho",
        area_ha=float(area) if area else 0.0,
        estagio=d.get("estagio") or "Crescimento",
        saude=d.get("saude") or "Bom",
        acoes=d.get("acoes") or ""
    )
    db.session.add(zona)
    db.session.commit()
    return jsonify(zona.to_dict()), 201


# ═══════════════════════════════════════════════════════════════
#  PUT /api/zones/<id>  — atualizar zona
# ═══════════════════════════════════════════════════════════════
@zones_export_bp.route("/api/zones/<int:zone_id>", methods=["PUT"])
@jwt_required()
def update_zone(zone_id):
    user = _get_user()
    zona = Zona.query.filter_by(id=zone_id, user_id=user.id).first()
    if not zona:
        return jsonify({"erro": "Zona não encontrada"}), 404

    d = request.get_json(silent=True) or {}
    if "nome"    in d: zona.nome    = (d["nome"] or "").strip()
    if "cultura" in d: zona.cultura = (d["cultura"] or "").strip()
    if "tipo"    in d: zona.tipo    = d["tipo"]
    if "area"    in d: zona.area_ha = float(d["area"] or 0)
    if "estagio" in d: zona.estagio = d["estagio"]
    if "saude"   in d: zona.saude   = d["saude"]
    if "acoes"   in d: zona.acoes   = d["acoes"]
    zona.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify(zona.to_dict()), 200


# ═══════════════════════════════════════════════════════════════
#  DELETE /api/zones/<id>  — eliminar zona
# ═══════════════════════════════════════════════════════════════
@zones_export_bp.route("/api/zones/<int:zone_id>", methods=["DELETE"])
@jwt_required()
def delete_zone(zone_id):
    user = _get_user()
    zona = Zona.query.filter_by(id=zone_id, user_id=user.id).first()
    if not zona:
        return jsonify({"erro": "Zona não encontrada"}), 404
    db.session.delete(zona)
    db.session.commit()
    return jsonify({"msg": "Zona eliminada"}), 200


# ═══════════════════════════════════════════════════════════════
#  EXPORTAÇÃO — helpers de dados
# ═══════════════════════════════════════════════════════════════
def _get_export_data(user, periodo, fazenda_id_filtro=None, is_admin=False):
    """Recolhe dados para exportação (XLSX ou PDF)."""
    dias_map = {"semanal": 7, "mensal": 30, "anual": 365}
    dias = dias_map.get(periodo, 30)
    cutoff = datetime.utcnow() - timedelta(days=dias)

    # Admin pode filtrar por fazenda; agricultor usa a sua
    fazenda = None
    device_ids = None

    if is_admin and fazenda_id_filtro:
        fazenda = Fazenda.query.get(int(fazenda_id_filtro))
        if fazenda:
            sensores = Sensor.query.filter_by(fazenda_id=fazenda.id).all()
            device_ids = [s.nome for s in sensores]
    elif not is_admin:
        fazenda = _get_fazenda(user)
        farm_activated_at = None
        if fazenda:
            farm_activated_at = fazenda.activated_at
            sensores = Sensor.query.filter_by(fazenda_id=fazenda.id).all()
            device_ids = [s.nome for s in sensores]

    q = DadosIoT.query.filter(DadosIoT.timestamp >= cutoff)
    if device_ids is not None:
        if device_ids:
            q = q.filter(DadosIoT.device_id.in_(device_ids))
        else:
            q = q.filter(False)

    rows = q.order_by(DadosIoT.timestamp.asc()).all()

    daily = {}
    for row in rows:
        day = row.timestamp.strftime("%Y-%m-%d")
        if day not in daily:
            daily[day] = {"temp": [], "hum_solo": [], "hum_ar": [], "pressao": [], "pragas": 0}
        if row.temperatura_ar is not None: daily[day]["temp"].append(row.temperatura_ar)
        if row.humidade_solo  is not None: daily[day]["hum_solo"].append(row.humidade_solo)
        if row.humidade_ar    is not None: daily[day]["hum_ar"].append(row.humidade_ar)
        if row.pressao_ar     is not None: daily[day]["pressao"].append(row.pressao_ar)
        if row.detecao_praga: daily[day]["pragas"] += 1

    def avg(lst): return round(sum(lst)/len(lst), 1) if lst else None

    chart_diario = [
        {"data": day,
         "temp": avg(v["temp"]),
         "hum_solo": avg(v["hum_solo"]),
         "hum_ar": avg(v["hum_ar"]),
         "pressao": avg(v["pressao"]),
         "pragas": v["pragas"]}
        for day, v in sorted(daily.items())
    ]

    all_temps  = [r.temperatura_ar for r in rows if r.temperatura_ar is not None]
    all_humid  = [r.humidade_solo  for r in rows if r.humidade_solo  is not None]
    all_hum_ar = [r.humidade_ar    for r in rows if r.humidade_ar    is not None]

    resumo = {
        "total_leituras": len(rows),
        "temp_media": avg(all_temps),
        "hum_solo_media": avg(all_humid),
        "hum_ar_media": avg(all_hum_ar),
        "pragas_detectadas": sum(1 for r in rows if r.detecao_praga),
        "temp_max": round(max(all_temps), 1) if all_temps else None,
        "temp_min": round(min(all_temps), 1) if all_temps else None,
    }

    # Zonas do agricultor
    zonas = []
    if not is_admin and user:
        zonas = Zona.query.filter_by(user_id=user.id).all()

    return {
        "fazenda": fazenda,
        "user": user,
        "periodo": periodo,
        "dias": dias,
        "resumo": resumo,
        "grafico_diario": chart_diario,
        "zonas": zonas,
        "gerado_em": datetime.utcnow(),
    }


# ═══════════════════════════════════════════════════════════════
#  GERAR XLSX
# ═══════════════════════════════════════════════════════════════
def _gerar_xlsx(data):
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.chart.series import SeriesLabel

    wb = Workbook()

    # ── Cores e estilos ──────────────────────────────────────────
    GREEN_DARK  = "1A5C38"
    GREEN_MED   = "2E8B57"
    GREEN_LIGHT = "D1FAE5"
    AMBER       = "F59E0B"
    BLUE        = "0EA5E9"
    RED         = "DC2626"
    GRAY_LIGHT  = "F9FAFB"
    GRAY_MED    = "E5E7EB"
    WHITE       = "FFFFFF"

    def hdr_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def thin_border():
        s = Side(style="thin", color="D1D5DB")
        return Border(left=s, right=s, top=s, bottom=s)

    def cell_style(ws, row, col, value, bold=False, color=WHITE, bg=None,
                   align="left", number_format=None, font_color="111827", font_size=10):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Calibri", bold=bold, color=font_color, size=font_size)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if bg:
            c.fill = hdr_fill(bg)
        c.border = thin_border()
        if number_format:
            c.number_format = number_format
        return c

    # ────────────────────────────────────────────────────────────
    # Folha 1 — RESUMO / CAPA
    # ────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumo"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 22
    ws1.column_dimensions["D"].width = 22
    ws1.row_dimensions[1].height = 50

    # Título principal (A1:D1)
    ws1.merge_cells("A1:D1")
    c = ws1.cell(row=1, column=1, value="🌿  AgroCaua — Relatório de Actividades")
    c.font = Font(name="Calibri", bold=True, color=WHITE, size=18)
    c.fill = hdr_fill(GREEN_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Sub-título com período
    periodo_label = {"semanal": "Últimos 7 dias", "mensal": "Últimos 30 dias",
                     "anual": "Últimos 365 dias"}.get(data["periodo"], data["periodo"])
    ws1.merge_cells("A2:D2")
    c2 = ws1.cell(row=2, column=1, value=f"Período: {periodo_label}  |  Gerado em: {data['gerado_em'].strftime('%d/%m/%Y às %H:%M')}")
    c2.font = Font(name="Calibri", color=WHITE, size=11, italic=True)
    c2.fill = hdr_fill(GREEN_MED)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 22

    # ── Info da Fazenda ──
    row = 4
    ws1.merge_cells(f"A{row}:D{row}")
    hdr = ws1.cell(row=row, column=1, value="INFORMAÇÕES DA FAZENDA")
    hdr.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    hdr.fill = hdr_fill(GREEN_MED)
    hdr.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[row].height = 20
    row += 1

    fazenda = data.get("fazenda")
    user    = data.get("user")
    info_fazenda = [
        ("Nome da Fazenda",  fazenda.nome if fazenda else "—"),
        ("Proprietário",     fazenda.proprietario if fazenda else (user.nome if user else "—")),
        ("Localização",      fazenda.localizacao if fazenda else "—"),
        ("Área Total (ha)",  f"{fazenda.hectares:.1f} ha" if fazenda and fazenda.hectares else "—"),
        ("Culturas",         ", ".join(fazenda.get_culturas_list()).title() if fazenda else "—"),
        ("Estado",           fazenda.status.capitalize() if fazenda else "—"),
    ]
    for label, val in info_fazenda:
        cell_style(ws1, row, 1, label, bold=True, bg=GRAY_LIGHT, font_color=GREEN_DARK)
        ws1.merge_cells(f"B{row}:D{row}")
        cell_style(ws1, row, 2, val, bg=WHITE)
        ws1.row_dimensions[row].height = 18
        row += 1

    # ── Info do Agricultor ──
    row += 1
    ws1.merge_cells(f"A{row}:D{row}")
    hdr2 = ws1.cell(row=row, column=1, value="DADOS DO AGRICULTOR")
    hdr2.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    hdr2.fill = hdr_fill(GREEN_MED)
    hdr2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[row].height = 20
    row += 1

    info_user = [
        ("Nome",  user.nome if user else "—"),
        ("Email", user.email if user else "—"),
        ("Perfil", user.role.capitalize() if user else "—"),
        ("Membro desde", user.created_at.strftime("%d/%m/%Y") if user and user.created_at else "—"),
    ]
    for label, val in info_user:
        cell_style(ws1, row, 1, label, bold=True, bg=GRAY_LIGHT, font_color=GREEN_DARK)
        ws1.merge_cells(f"B{row}:D{row}")
        cell_style(ws1, row, 2, val, bg=WHITE)
        ws1.row_dimensions[row].height = 18
        row += 1

    # ── KPIs do Resumo ──
    row += 1
    ws1.merge_cells(f"A{row}:D{row}")
    hdr3 = ws1.cell(row=row, column=1, value="RESUMO DO PERÍODO")
    hdr3.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    hdr3.fill = hdr_fill(GREEN_DARK)
    hdr3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[row].height = 20
    row += 1

    r = data["resumo"]
    kpis = [
        ("Total de Leituras",           r.get("total_leituras", "—")),
        ("Temperatura Média (°C)",       r.get("temp_media", "—")),
        ("Temperatura Máxima (°C)",      r.get("temp_max", "—")),
        ("Temperatura Mínima (°C)",      r.get("temp_min", "—")),
        ("Humidade Solo Média (%)",      r.get("hum_solo_media", "—")),
        ("Humidade Ar Média (%)",        r.get("hum_ar_media", "—")),
        ("Pragas Detectadas",            r.get("pragas_detectadas", 0)),
    ]
    for i, (label, val) in enumerate(kpis):
        bg = GRAY_LIGHT if i % 2 == 0 else WHITE
        cell_style(ws1, row, 1, label, bold=True, bg=bg, font_color="374151")
        ws1.merge_cells(f"B{row}:D{row}")
        cell_style(ws1, row, 2, val, bg=bg, align="center",
                   font_color=(RED if label == "Pragas Detectadas" and val else "111827"),
                   bold=(label == "Pragas Detectadas" and bool(val)))
        ws1.row_dimensions[row].height = 18
        row += 1

    # ────────────────────────────────────────────────────────────
    # Folha 2 — DADOS DIÁRIOS
    # ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Dados Diários")
    ws2.sheet_view.showGridLines = False

    cols_width = {"A": 14, "B": 16, "C": 18, "D": 16, "E": 18, "F": 14, "G": 14}
    for col, width in cols_width.items():
        ws2.column_dimensions[col].width = width

    # Título
    ws2.merge_cells("A1:G1")
    t = ws2.cell(row=1, column=1, value="Dados Diários de Sensores IoT")
    t.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    t.fill = hdr_fill(GREEN_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 36

    ws2.merge_cells("A2:G2")
    st = ws2.cell(row=2, column=1, value=f"Período: {periodo_label}  |  Fazenda: {fazenda.nome if fazenda else '—'}")
    st.font = Font(italic=True, color="6B7280", size=10, name="Calibri")
    st.alignment = Alignment(horizontal="center")
    ws2.row_dimensions[2].height = 16

    # Cabeçalhos da tabela
    headers = ["Data", "Temp. Ar (°C)", "Hum. Solo (%)", "Hum. Ar (%)", "Pressão (hPa)", "Pragas", "Estado"]
    for col_i, h in enumerate(headers, 1):
        c = ws2.cell(row=4, column=col_i, value=h)
        c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
        c.fill = hdr_fill(GREEN_MED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    ws2.row_dimensions[4].height = 22

    # Dados
    for row_i, d_row in enumerate(data["grafico_diario"], 5):
        bg = GRAY_LIGHT if row_i % 2 == 0 else WHITE

        pragas = d_row.get("pragas", 0) or 0
        temp   = d_row.get("temp")
        hum_s  = d_row.get("hum_solo")

        # Estado calculado
        if pragas > 0:
            estado = "⚠ Alerta"
            est_color = RED
        elif (temp is not None and temp > 35) or (hum_s is not None and hum_s < 20):
            estado = "△ Atenção"
            est_color = AMBER
        else:
            estado = "✓ Normal"
            est_color = "15803D"

        cell_style(ws2, row_i, 1, d_row.get("data", ""), bg=bg, align="center")
        cell_style(ws2, row_i, 2, temp,  bg=bg, align="center",
                   font_color=(RED if temp and temp > 35 else ("3B82F6" if temp and temp < 15 else "111827")))
        cell_style(ws2, row_i, 3, hum_s, bg=bg, align="center",
                   font_color=(RED if hum_s and hum_s < 20 else ("3B82F6" if hum_s and hum_s > 80 else "111827")))
        cell_style(ws2, row_i, 4, d_row.get("hum_ar"),   bg=bg, align="center")
        cell_style(ws2, row_i, 5, d_row.get("pressao"),  bg=bg, align="center")
        cell_style(ws2, row_i, 6, pragas, bg=bg, align="center",
                   bold=(pragas > 0), font_color=(RED if pragas > 0 else "111827"))
        cell_style(ws2, row_i, 7, estado, bg=bg, align="center",
                   bold=True, font_color=est_color)
        ws2.row_dimensions[row_i].height = 16

    # ── Gráfico de Temperatura ──
    if len(data["grafico_diario"]) > 1:
        n_rows = len(data["grafico_diario"])
        data_ref  = Reference(ws2, min_col=1, min_row=5, max_row=4 + n_rows)
        temp_ref  = Reference(ws2, min_col=2, min_row=4, max_row=4 + n_rows)
        hum_ref   = Reference(ws2, min_col=3, min_row=4, max_row=4 + n_rows)

        chart = LineChart()
        chart.title = "Temperatura e Humidade do Solo"
        chart.style = 10
        chart.height = 12
        chart.width  = 22
        chart.y_axis.title = "Valor"
        chart.x_axis.title = "Data"

        s1 = chart.series.append
        from openpyxl.chart import Series
        serie_t = Series(temp_ref, title="Temperatura (°C)")
        serie_t.graphicalProperties.line.solidFill = "F59E0B"
        serie_t.graphicalProperties.line.width = 20000
        chart.series.append(serie_t)

        serie_h = Series(hum_ref, title="Hum. Solo (%)")
        serie_h.graphicalProperties.line.solidFill = "0EA5E9"
        serie_h.graphicalProperties.line.width = 20000
        chart.series.append(serie_h)

        chart.set_categories(data_ref)
        ws2.add_chart(chart, f"A{6 + n_rows + 2}")

    # ────────────────────────────────────────────────────────────
    # Folha 3 — ZONAS DE CULTIVO
    # ────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Zonas de Cultivo")
    ws3.sheet_view.showGridLines = False

    for col, width in {"A": 22, "B": 18, "C": 16, "D": 14, "E": 18, "F": 18, "G": 35}.items():
        ws3.column_dimensions[col].width = width

    ws3.merge_cells("A1:G1")
    t3 = ws3.cell(row=1, column=1, value="Zonas de Cultivo")
    t3.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    t3.fill = hdr_fill(GREEN_DARK)
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 36

    h3 = ["Zona / Nome", "Cultura", "Tipo", "Área (ha)", "Estágio", "Saúde", "Acções / Notas"]
    for ci, h in enumerate(h3, 1):
        c = ws3.cell(row=3, column=ci, value=h)
        c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
        c.fill = hdr_fill(GREEN_MED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    ws3.row_dimensions[3].height = 22

    zonas = data.get("zonas", [])
    if zonas:
        saude_color = {"Excelente": "15803D", "Bom": "0284C7", "Atenção": "B45309", "Crítico": RED}
        for ri, z in enumerate(zonas, 4):
            bg = GRAY_LIGHT if ri % 2 == 0 else WHITE
            saude = z.saude or "—"
            sc = saude_color.get(saude, "374151")
            cell_style(ws3, ri, 1, z.nome,    bg=bg, bold=True, font_color=GREEN_DARK)
            cell_style(ws3, ri, 2, z.cultura,  bg=bg)
            cell_style(ws3, ri, 3, z.tipo,     bg=bg)
            cell_style(ws3, ri, 4, z.area_ha,  bg=bg, align="center", number_format="0.00")
            cell_style(ws3, ri, 5, z.estagio,  bg=bg, align="center")
            cell_style(ws3, ri, 6, saude,      bg=bg, align="center", bold=True, font_color=sc)
            cell_style(ws3, ri, 7, z.acoes,    bg=bg)
            ws3.row_dimensions[ri].height = 18
    else:
        ws3.merge_cells("A4:G4")
        nc = ws3.cell(row=4, column=1, value="Sem zonas de cultivo registadas.")
        nc.font = Font(italic=True, color="9CA3AF", name="Calibri")
        nc.alignment = Alignment(horizontal="center")

    # ────────────────────────────────────────────────────────────
    # Folha 4 — ESTATÍSTICAS (se tiver dados suficientes)
    # ────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Estatísticas")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 18

    ws4.merge_cells("A1:B1")
    t4 = ws4.cell(row=1, column=1, value="Estatísticas de Sensores")
    t4.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    t4.fill = hdr_fill(GREEN_DARK)
    t4.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 36

    stats = [
        ("Indicador", "Valor", True),
        ("Total de leituras IoT",    r.get("total_leituras", 0), False),
        ("Temperatura média (°C)",   r.get("temp_media", "—"), False),
        ("Temperatura máxima (°C)",  r.get("temp_max", "—"), False),
        ("Temperatura mínima (°C)",  r.get("temp_min", "—"), False),
        ("Humidade solo média (%)",  r.get("hum_solo_media", "—"), False),
        ("Humidade ar média (%)",    r.get("hum_ar_media", "—"), False),
        ("Total pragas detectadas",  r.get("pragas_detectadas", 0), False),
        ("Dias com dados",           len(data["grafico_diario"]), False),
    ]
    for si, (label, val, is_hdr) in enumerate(stats, 3):
        bg_l = GREEN_MED if is_hdr else (GRAY_LIGHT if si % 2 == 0 else WHITE)
        fc_l = WHITE if is_hdr else GREEN_DARK
        cell_style(ws4, si, 1, label, bold=is_hdr, bg=bg_l, font_color=fc_l)
        cell_style(ws4, si, 2, val,   bold=is_hdr, bg=bg_l, font_color=fc_l, align="center")
        ws4.row_dimensions[si].height = 18

    # Salvar em memória
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════
#  GERAR PDF
# ═══════════════════════════════════════════════════════════════
def _gerar_pdf(data):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                     TableStyle, HRFlowable, PageBreak)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.8*cm, rightMargin=1.8*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    # ── Cores ────────────────────────────────────────────────────
    C_GREEN_DARK  = colors.HexColor("#1A5C38")
    C_GREEN_MED   = colors.HexColor("#2E8B57")
    C_GREEN_LIGHT = colors.HexColor("#D1FAE5")
    C_AMBER       = colors.HexColor("#F59E0B")
    C_RED         = colors.HexColor("#DC2626")
    C_BLUE        = colors.HexColor("#0EA5E9")
    C_GRAY_LIGHT  = colors.HexColor("#F9FAFB")
    C_GRAY_MED    = colors.HexColor("#E5E7EB")
    C_TEXT        = colors.HexColor("#111827")
    C_TEXT_LIGHT  = colors.HexColor("#6B7280")
    C_WHITE       = colors.white

    # ── Estilos ──────────────────────────────────────────────────
    styles = getSampleStyleSheet()
    def ps(name, **kw):
        return ParagraphStyle(name, **kw)

    style_h1 = ps("H1", fontSize=20, fontName="Helvetica-Bold",
                   textColor=C_WHITE, alignment=TA_CENTER, leading=26)
    style_h2 = ps("H2", fontSize=13, fontName="Helvetica-Bold",
                   textColor=C_WHITE, alignment=TA_LEFT, leading=18)
    style_h3 = ps("H3", fontSize=11, fontName="Helvetica-Bold",
                   textColor=C_GREEN_DARK, alignment=TA_LEFT, leading=16)
    style_body = ps("Body", fontSize=9, fontName="Helvetica",
                    textColor=C_TEXT, leading=13)
    style_small = ps("Small", fontSize=8, fontName="Helvetica",
                     textColor=C_TEXT_LIGHT, leading=12)
    style_center = ps("Center", fontSize=9, fontName="Helvetica",
                      textColor=C_TEXT, alignment=TA_CENTER, leading=13)

    def tbl_style_base():
        return TableStyle([
            ("BACKGROUND",   (0,0), (-1,0), C_GREEN_MED),
            ("TEXTCOLOR",    (0,0), (-1,0), C_WHITE),
            ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,0), 9),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [C_GRAY_LIGHT, C_WHITE]),
            ("FONTNAME",     (0,1), (-1,-1), "Helvetica"),
            ("FONTSIZE",     (0,1), (-1,-1), 8),
            ("TEXTCOLOR",    (0,1), (-1,-1), C_TEXT),
            ("GRID",         (0,0), (-1,-1), 0.5, C_GRAY_MED),
            ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING",   (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0), (-1,-1), 5),
            ("LEFTPADDING",  (0,0), (-1,-1), 6),
            ("RIGHTPADDING", (0,0), (-1,-1), 6),
            ("ROWBACKGROUNDS", (0,0), (-1,0), [C_GREEN_MED]),
            ("ALIGN",        (1,1), (-1,-1), "CENTER"),
        ])

    # ── Cabeçalho colorido ───────────────────────────────────────
    def header_block(title, subtitle=""):
        tbl = Table([[Paragraph(title, style_h1)]], colWidths=[doc.width])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), C_GREEN_DARK),
            ("TOPPADDING",    (0,0), (-1,-1), 10),
            ("BOTTOMPADDING", (0,0), (-1,-1), 10),
        ]))
        elems = [tbl]
        if subtitle:
            st = Table([[Paragraph(subtitle, ps("Sub", fontSize=10,
                                                fontName="Helvetica-Oblique",
                                                textColor=C_WHITE,
                                                alignment=TA_CENTER))]], colWidths=[doc.width])
            st.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,-1), C_GREEN_MED),
                ("TOPPADDING", (0,0), (-1,-1), 4),
                ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ]))
            elems.append(st)
        return elems

    def section_header(text):
        tbl = Table([[Paragraph(text, style_h2)]], colWidths=[doc.width])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), C_GREEN_MED),
            ("TOPPADDING",    (0,0), (-1,-1), 6),
            ("BOTTOMPADDING", (0,0), (-1,-1), 6),
            ("LEFTPADDING",   (0,0), (-1,-1), 8),
        ]))
        return tbl

    # ════════════════════════════════════════════════════════════
    story = []
    periodo_label = {"semanal": "Semanal — Últimos 7 dias",
                     "mensal":  "Mensal — Últimos 30 dias",
                     "anual":   "Anual — Últimos 365 dias"}.get(data["periodo"], data["periodo"])

    # ── CAPA ────────────────────────────────────────────────────
    story += header_block("🌿  AgroCaua", f"Relatório de Actividades — {periodo_label}")
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(f"Gerado em: {data['gerado_em'].strftime('%d/%m/%Y às %H:%M UTC')}",
                            style_small))
    story.append(Spacer(1, 0.6*cm))

    # ── INFO FAZENDA ─────────────────────────────────────────────
    story.append(section_header("INFORMAÇÕES DA FAZENDA"))
    story.append(Spacer(1, 0.2*cm))

    fazenda = data.get("fazenda")
    user    = data.get("user")

    faz_data = [
        ["Campo", "Valor"],
        ["Nome da Fazenda",   fazenda.nome if fazenda else "—"],
        ["Proprietário",      fazenda.proprietario if fazenda else (user.nome if user else "—")],
        ["Localização",       fazenda.localizacao if fazenda else "—"],
        ["Área Total (ha)",   f"{fazenda.hectares:.1f} ha" if fazenda and fazenda.hectares else "—"],
        ["Culturas",          ", ".join(fazenda.get_culturas_list()).title() if fazenda else "—"],
        ["Estado",            fazenda.status.capitalize() if fazenda else "—"],
    ]
    faz_tbl = Table(faz_data, colWidths=[6*cm, doc.width - 6*cm])
    faz_tbl.setStyle(tbl_style_base())
    story.append(faz_tbl)
    story.append(Spacer(1, 0.5*cm))

    # ── INFO AGRICULTOR ──────────────────────────────────────────
    story.append(section_header("DADOS DO AGRICULTOR"))
    story.append(Spacer(1, 0.2*cm))

    usr_data = [
        ["Campo", "Valor"],
        ["Nome",          user.nome if user else "—"],
        ["Email",         user.email if user else "—"],
        ["Perfil",        user.role.capitalize() if user else "—"],
        ["Membro desde",  user.created_at.strftime("%d/%m/%Y") if user and user.created_at else "—"],
    ]
    usr_tbl = Table(usr_data, colWidths=[6*cm, doc.width - 6*cm])
    usr_tbl.setStyle(tbl_style_base())
    story.append(usr_tbl)
    story.append(Spacer(1, 0.5*cm))

    # ── RESUMO KPIs ──────────────────────────────────────────────
    story.append(section_header("RESUMO DO PERÍODO"))
    story.append(Spacer(1, 0.2*cm))

    r = data["resumo"]
    kpi_data = [
        ["Indicador", "Valor"],
        ["Total de Leituras IoT",     str(r.get("total_leituras", 0))],
        ["Temperatura Média",          f"{r['temp_media']} °C" if r.get("temp_media") is not None else "—"],
        ["Temperatura Máxima",         f"{r['temp_max']} °C"   if r.get("temp_max")   is not None else "—"],
        ["Temperatura Mínima",         f"{r['temp_min']} °C"   if r.get("temp_min")   is not None else "—"],
        ["Humidade Solo Média",        f"{r['hum_solo_media']} %" if r.get("hum_solo_media") is not None else "—"],
        ["Humidade Ar Média",          f"{r['hum_ar_media']} %"   if r.get("hum_ar_media")   is not None else "—"],
        ["Pragas Detectadas",          str(r.get("pragas_detectadas", 0))],
    ]
    kpi_tbl = Table(kpi_data, colWidths=[9*cm, doc.width - 9*cm])
    kpi_style = tbl_style_base()
    pragas_val = r.get("pragas_detectadas", 0)
    if pragas_val:
        kpi_style.add("TEXTCOLOR", (1, len(kpi_data)-1), (1, len(kpi_data)-1), C_RED)
        kpi_style.add("FONTNAME",  (1, len(kpi_data)-1), (1, len(kpi_data)-1), "Helvetica-Bold")
    kpi_tbl.setStyle(kpi_style)
    story.append(kpi_tbl)

    # ── DADOS DIÁRIOS ────────────────────────────────────────────
    story.append(PageBreak())
    story += header_block("Dados Diários de Sensores IoT")
    story.append(Spacer(1, 0.4*cm))

    dias_data = [["Data", "Temp (°C)", "Hum Solo (%)", "Hum Ar (%)", "Pressão (hPa)", "Pragas", "Estado"]]
    for d_row in data["grafico_diario"]:
        pragas = d_row.get("pragas", 0) or 0
        temp   = d_row.get("temp")
        hum_s  = d_row.get("hum_solo")
        if pragas > 0:
            estado = "⚠ Alerta"
        elif (temp is not None and temp > 35) or (hum_s is not None and hum_s < 20):
            estado = "△ Atenção"
        else:
            estado = "✓ Normal"
        dias_data.append([
            d_row.get("data", "—"),
            f"{temp}" if temp is not None else "—",
            f"{hum_s}" if hum_s is not None else "—",
            f"{d_row.get('hum_ar')}" if d_row.get("hum_ar") is not None else "—",
            f"{d_row.get('pressao')}" if d_row.get("pressao") is not None else "—",
            str(pragas),
            estado
        ])

    col_ws = [2.8*cm, 2.2*cm, 2.5*cm, 2.2*cm, 2.8*cm, 1.8*cm, 2.5*cm]
    if not data["grafico_diario"]:
        dias_data.append(["Sem dados para este período"] + [""] * 6)

    dias_tbl = Table(dias_data, colWidths=col_ws, repeatRows=1)
    days_style = tbl_style_base()
    # Colorir estado
    for ri in range(1, len(dias_data)):
        estado_val = dias_data[ri][6]
        if "Alerta" in str(estado_val):
            days_style.add("TEXTCOLOR", (6, ri), (6, ri), C_RED)
            days_style.add("FONTNAME",  (6, ri), (6, ri), "Helvetica-Bold")
        elif "Atenção" in str(estado_val):
            days_style.add("TEXTCOLOR", (6, ri), (6, ri), C_AMBER)
            days_style.add("FONTNAME",  (6, ri), (6, ri), "Helvetica-Bold")
        else:
            days_style.add("TEXTCOLOR", (6, ri), (6, ri), colors.HexColor("#15803D"))
    dias_tbl.setStyle(days_style)
    story.append(dias_tbl)

    # ── ZONAS DE CULTIVO ─────────────────────────────────────────
    zonas = data.get("zonas", [])
    if zonas:
        story.append(PageBreak())
        story += header_block("Zonas de Cultivo")
        story.append(Spacer(1, 0.4*cm))

        z_data = [["Zona", "Cultura", "Tipo", "Área (ha)", "Estágio", "Saúde"]]
        saude_colors = {"Excelente": colors.HexColor("#15803D"),
                        "Bom": C_BLUE, "Atenção": C_AMBER, "Crítico": C_RED}
        for z in zonas:
            z_data.append([
                z.nome, z.cultura, z.tipo or "—",
                f"{z.area_ha:.2f} ha" if z.area_ha else "—",
                z.estagio or "—", z.saude or "—"
            ])

        z_col_ws = [3.8*cm, 3*cm, 2.5*cm, 2.2*cm, 2.5*cm, 2.5*cm]
        z_tbl = Table(z_data, colWidths=z_col_ws, repeatRows=1)
        z_style = tbl_style_base()
        for ri in range(1, len(z_data)):
            saude = z_data[ri][5]
            sc = saude_colors.get(saude, C_TEXT)
            z_style.add("TEXTCOLOR", (5, ri), (5, ri), sc)
            z_style.add("FONTNAME",  (5, ri), (5, ri), "Helvetica-Bold")
        z_tbl.setStyle(z_style)
        story.append(z_tbl)

        # Notas/Acções por zona
        story.append(Spacer(1, 0.4*cm))
        story.append(Paragraph("Notas por Zona", style_h3))
        story.append(Spacer(1, 0.2*cm))
        for z in zonas:
            if z.acoes:
                story.append(Paragraph(f"<b>{z.nome}:</b> {z.acoes}", style_body))
                story.append(Spacer(1, 0.15*cm))

    # ── Rodapé ───────────────────────────────────────────────────
    story.append(Spacer(1, 1*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=C_GRAY_MED))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(
        f"Relatório gerado automaticamente pela plataforma AgroCaua — {data['gerado_em'].strftime('%d/%m/%Y')}",
        ps("Footer", fontSize=7.5, fontName="Helvetica-Oblique",
           textColor=C_TEXT_LIGHT, alignment=TA_CENTER)
    ))

    doc.build(story)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════
#  POST /api/relatorios/exportar
# ═══════════════════════════════════════════════════════════════
@zones_export_bp.route("/api/relatorios/exportar", methods=["POST"])
@jwt_required()
def exportar_relatorio():
    user = _get_user()
    if not user:
        return jsonify({"erro": "Utilizador não encontrado"}), 404

    d = request.get_json(silent=True) or {}
    formato     = (d.get("formato") or "xlsx").lower()
    periodo     = d.get("periodo", "mensal")
    fazenda_id  = d.get("fazenda_id")
    is_admin    = user.role in ("admin", "superadmin")

    if formato not in ("xlsx", "pdf"):
        return jsonify({"erro": "Formato inválido. Use 'xlsx' ou 'pdf'"}), 400

    # Recolher dados
    data = _get_export_data(user, periodo, fazenda_id_filtro=fazenda_id, is_admin=is_admin)

    # Nome do ficheiro
    now_str = datetime.utcnow().strftime("%Y%m%d_%H%M")
    periodo_slug = {"semanal": "7d", "mensal": "30d", "anual": "365d"}.get(periodo, periodo)
    filename = f"agrocaua_relatorio_{periodo_slug}_{now_str}.{formato}"

    if formato == "xlsx":
        buf = _gerar_xlsx(data)
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        buf = _gerar_pdf(data)
        mime = "application/pdf"

    return send_file(
        buf,
        mimetype=mime,
        as_attachment=True,
        download_name=filename
    )
